


from openpyxl import Workbook,load_workbook
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
from cycler import cycler
import math
from scipy.interpolate import interp1d, CubicSpline


def find_insert_position(arr, target):
    """
    使用二分查找算法查找目标数字在升序数组中的插入位置。
    如果目标数字已经在数组中，返回其索引；如果目标数字不在数组中，
    返回目标数字应该插入的位置，使得插入后数组仍然保持升序。

    :param arr: 升序数组
    :param target: 要查找的目标数字
    :return: 目标数字在数组中的插入位置
    """
    left, right = 0, len(arr) - 1

    while left <= right:
        mid = (left + right) // 2  # 计算中间索引

        if arr[mid] == target:
            return mid  # 找到目标数字，返回其索引
        elif arr[mid] < target:
            left = mid + 1  # 目标在右侧子数组
        else:
            right = mid - 1  # 目标在左侧子数组

    # 如果未找到目标数字，left 的位置即为插入位置
    return left

def sort(wavelength_1d,spectrum_1d,var_1d):
    wavelength_1d=np.asarray(wavelength_1d)
    spectrum_1d=np.asarray(spectrum_1d)
    var_1d=np.asarray(var_1d)

    idx = np.argsort(wavelength_1d)

    # 2. 重新排序三个数组
    wavelength_sorted = wavelength_1d[idx]
    spectrum_sorted   = spectrum_1d[idx]
    var_sorted        = var_1d[idx]

    return wavelength_sorted, spectrum_sorted, var_sorted


def get_txt(file_path):
    try:
        wavelength, flux, error = np.loadtxt(file_path,
                                     comments='#',      # 跳过所有 # 开头的行
                                     unpack=True)       # 直接拆成三列数组
        return wavelength[find_insert_position(wavelength,1.0):find_insert_position(wavelength,2.5)],flux[find_insert_position(wavelength,1.0):find_insert_position(wavelength,2.5)],error[find_insert_position(wavelength,1.0):find_insert_position(wavelength,2.5)]
    except:
        wavelength, flux = np.loadtxt(file_path,
                                     comments='#',      # 跳过所有 # 开头的行
                           unpack=True)       # 直接拆成2列数组
        wavelength=wavelength/1e4
        e=flux
        return wavelength[find_insert_position(wavelength,1.0):find_insert_position(wavelength,2.5)],flux[find_insert_position(wavelength,1.0):find_insert_position(wavelength,2.5)],e[find_insert_position(wavelength,1.0):find_insert_position(wavelength,2.5)]


def extract_values_from_txt(file_path):
    """
    从指定的txt文件中提取teff、logg、meta和alpha的数值。
    
    参数:
        file_path (str): txt文件的路径。
    
    返回:
        dict: 包含teff、logg、meta和alpha的字典。
    """
    # 初始化一个字典来存储提取的数值
    values = {
        "teff": None,
        "logg": None,
        "meta": None,
        "alpha": None
    }
    
    try:
        # 打开文件并逐行读取
        with open(file_path, "r") as file:
            for line in file:
                # 去掉行首和行尾的空白字符
                line = line.strip()
                # 检查是否包含teff
                if "teff" in line:
                    # 提取teff的数值
                    values["teff"] = float(line.split("=")[1].split("K")[0].strip())
                # 检查是否包含logg
                elif "logg" in line:
                    # 提取logg的数值
                    values["logg"] = float(line.split("=")[1].split("log")[0].strip())
                # 检查是否包含meta
                elif "meta" in line:
                    # 提取meta的数值
                    values["meta"] = float(line.split("=")[1].split('(')[0].strip())
                # 检查是否包含alpha
                elif "alpha" in line:
                    # 提取alpha的数值
                    values["alpha"] = float(line.split("=")[1].split('(')[0].strip())
    except FileNotFoundError:
        print(f"文件 {file_path} 未找到。")
    except Exception as e:
        print(f"读取文件时发生错误：{e}")
    
    return values


def excel_and_add_row(file_path, data):
    """
    创建一个Excel文件，并在表格末尾增加一行信息。
    
    参数:
        file_path (str): Excel文件的路径。
        data (list): 要添加到表格末尾的行数据。
    """
    # 加载现有的工作簿
    workbook = load_workbook(file_path)
    # 获取默认的工作表
    sheet = workbook.active
    
    # 在表格末尾添加一行数据
    sheet.append(data)
    
    # 保存工作簿到指定路径
    workbook.save(file_path)
    print(f"Excel文件已创建并添加了一行数据：{file_path}")

def create_excel_and_add_row(file_path, data):
    """
    创建一个Excel文件，并在表格末尾增加一行信息。
    
    参数:
        file_path (str): Excel文件的路径。
        data (list): 要添加到表格末尾的行数据。
    """
    # 创建一个新的工作簿
    workbook = Workbook()
    # 获取默认的工作表
    sheet = workbook.active
    
    # 在表格末尾添加一行数据
    sheet.append(data)
    
    # 保存工作簿到指定路径
    workbook.save(file_path)
    print(f"Excel文件已创建并添加了一行数据：{file_path}")


def get_Ck(wave,flux,wave_K,flux_K,error_K,wave_H,flux_H,error_H,wave_model_H,flux_model_H):

    #left_index=0
    # #print("left_index",left_index)
    # right_index=find_insert_position(wave,wave_K[-1])

    '''
    sum1 = np.sum( flux * flux_K / err_K**2 ) + np.sum( flux * flux_H / err_H**2 )
    sum2 = np.sum( flux**2 / err_K**2 ) + np.sum( flux**2 / err**2 )#有问题

    sum1 = np.sum( flux_model_K * flux_K / err_K**2 ) + np.sum( flux_model_H * flux_H / err_H**2 )
    sum2 = np.sum( flux_model_K**2 / err_K**2 ) + np.sum( flux_model_H**2 / err_H**2 )

    '''

    sum1=0
    sum2=0
    for i in range(0,len(wave_K)):
        #print(i,left_index+i,wave_K[i],wave[left_index+i])
        #break
        sum1=sum1+flux[i]*flux_K[i]/(error_K[i]**2)
        sum2=sum2+flux[i]**2/(error_K[i]**2)

    for i in range(0,len(wave_H)):
        sum1=sum1+flux_model_H[i]*flux_H[i]/(error_H[i]**2)
        sum2=sum2+flux_model_H[i]**2/(error_H[i]**2)

    Ck=sum1/sum2
    #Ck=sum(wave[left_index+i]*wave_K[i]/(error_K[i]**2))/sum(wave[left_index+i]**2/(error_K[i]**2))
    print('Ck:',Ck)
    return Ck

def get_Gk(wave,flux,wave_K,flux_K,error_K,Ck,wave_H,flux_H,error_H,wave_model_H,flux_model_H):

    # left_index=0
    # #print("left_index",left_index)
    # right_index=find_insert_position(wave,wave_K[-1])

    sum1=0
    
    for i in range(0,len(wave_K)):
        #print(i,left_index+i,wave_K[i],wave[left_index+i])
        #break
        sum1=sum1+(flux_K[i]-Ck*flux[i])**2/(error_K[i]**2)

    for i in range(0,len(wave_H)):
        sum1=sum1+(flux_H[i]-Ck*flux_model_H[i])**2/(error_H[i]**2)

    Gk=sum1
    #Ck=sum(wave[left_index+i]*wave_K[i]/(error_K[i]**2))/sum(wave[left_index+i]**2/(error_K[i]**2))
    print('Gk:',Gk)
    return Gk
    

if __name__=="__main__":
    # file_path = "code/part_2/data_for_part2/bt_settle_cut/bt-settl_131.dat_cut.txt"
    # values = extract_values_from_txt(file_path)
    # print(values)  # 输出提取的数值

    K_path='code/part_2/data_for_part2/K.txt'
    H_path='code/part_2/data_for_part2/H.txt'
    wavelength_K, flux_K, error_K = get_txt(K_path)
    wavelength_K, flux_K, error_K = sort(wavelength_K, flux_K, error_K)
    wavelength_H, flux_H, error_H = get_txt(H_path)
    wavelength_H, flux_H, error_H = sort(wavelength_H, flux_H, error_H)


    # 示例用法
    excel_path = "code/part_2/outcome_for_part2/example.xlsx"  # 替换为你的Excel文件路径
    data = ["number", "teff", "logg","meta","alpha","Ck","Gk"]  # 替换为你想要添加的行数据
    create_excel_and_add_row(excel_path, data)
    number=1
    input_dir = 'models/bt-settle_K' # 替换为你的文件夹路径
    for file_name in os.listdir(input_dir):
        if file_name.endswith('.txt'):
            file_path_K = os.path.join(input_dir, file_name)
            print(file_path_K)
            file_path_H = os.path.join(input_dir.replace('K', 'H'), file_name)
            print(file_path_H)
            #print(file_path)
            values = extract_values_from_txt(file_path_K)
            print(values)  # 输出提取的数值
            wl,flr,aaaaaa=get_txt(file_path_K)
            wl,flr,aaaaaa=sort(wl,flr,aaaaaa)

            wl_H, flr_H, aaaaaa_H = get_txt(file_path_H)
            wl_H, flr_H, aaaaaa_H = sort(wl_H, flr_H, aaaaaa_H)

            #flr=np.zeros_like(flr) +1e-13   ###二次检查部分
            try:
                Ck=get_Ck(wl,flr,wavelength_K,flux_K,error_K,wave_H=wavelength_H,flux_H=flux_H,error_H=error_H,wave_model_H=wl_H,flux_model_H=flr_H)
                Gk=get_Gk(wl,flr,wavelength_K,flux_K,error_K,Ck,wave_H=wavelength_H,flux_H=flux_H,error_H=error_H,wave_model_H=wl_H,flux_model_H=flr_H)

                #break        ###二次检查部分

                excel_and_add_row(excel_path, [number, values["teff"], values["logg"], values["meta"], values["alpha"], Ck, Gk])
                number=number+1    
            except:
                print("出错了")
            #break    
