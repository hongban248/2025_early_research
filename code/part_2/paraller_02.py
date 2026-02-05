#本文件先cut原始的bt-settle模型，然后进行vr，vsini处理，最后插值到观测波段
#本脚本绘制全部的ck，gk图像，经过vr，vsini处理后的结果
#本脚本使用了air_to_vacuum函数将模型从空气波长转换为真空波长，原始数据BC处理了
#本脚本使用c的动态链接库以及多线程处理

#尝试这个降分变率
import coronagraph as cg  #这个包还有疑似神秘小bug，需要修改一行代码才能调用
#print(cg.__version__)
import coronagraph as cg
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
from cycler import cycler
import math
from scipy.interpolate import interp1d, CubicSpline
from openpyxl import Workbook,load_workbook
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
import ctypes
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed


def sort(wavelength_1d,spectrum_1d,var_1d):
    wavelength_1d=np.asarray(wavelength_1d)
    spectrum_1d=np.asarray(spectrum_1d)
    var_1d=np.asarray(var_1d)

    idx = np.argsort(wavelength_1d)

    # 2. 重新排序三个数组
    wavelength_sorted = wavelength_1d[idx]
    spectrum_sorted   = spectrum_1d[idx]
    var_sorted        = var_1d[idx]

    return wavelength_sorted, spectrum_sorted, var_sorted


lib_path='code/part_2/func_export.so'
# 2. 加载
lib = ctypes.CDLL(str(lib_path))

def echo_a_fromc():
    lib.echo_a_export.argtypes=None
    lib.echo_a_export.restype=None

    lib.echo_a_export()

def sort_c(wave, flux, error):
    # 统一转成 float64 的可写 C 连续数组
    wave = np.asarray(wave, dtype=np.float64, order='C')
    flux = np.asarray(flux, dtype=np.float64, order='C')
    error = np.asarray(error, dtype=np.float64, order='C')
    n = wave.size
    if not (n == flux.size == error.size):
        raise ValueError('三个数组长度不一致')

    # 保证可写
    wave = np.require(wave, requirements='CW')
    flux = np.require(flux, requirements='CW')
    error = np.require(error, requirements='CW')

    # 设置函数签名
    lib.sort_export.argtypes = (
        ctypes.POINTER(ctypes.c_double),
        ctypes.POINTER(ctypes.c_double),
        ctypes.POINTER(ctypes.c_double),
        ctypes.c_int)          # ← 不是 POINTER，只是值
    lib.sort_export.restype = None

    # 调用
    lib.sort_export(
        wave.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        flux.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        error.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        ctypes.c_int(n))

    return wave, flux, error
def get_txt(file_path):
    try:
        wavelength, flux, error = np.loadtxt(file_path,
                                     comments='#',      # 跳过所有 # 开头的行
                                     unpack=True)       # 直接拆成三列数组
        return wavelength[find_insert_position(wavelength,0.8):find_insert_position(wavelength,2.5)],flux[find_insert_position(wavelength,0.8):find_insert_position(wavelength,2.5)],error[find_insert_position(wavelength,0.8):find_insert_position(wavelength,2.5)]
    except:
        wavelength, flux = np.loadtxt(file_path,
                                     comments='#',      # 跳过所有 # 开头的行
                           unpack=True)       # 直接拆成2列数组
        wavelength=wavelength/1e4
        e=flux
        return wavelength[find_insert_position(wavelength,0.8):find_insert_position(wavelength,2.5)],flux[find_insert_position(wavelength,0.8):find_insert_position(wavelength,2.5)],e[find_insert_position(wavelength,0.8):find_insert_position(wavelength,2.5)]
def find_insert_position(arr, target):
    """
    使用二分查找算法查找目标数字在升序数组中的插入位置。
    如果目标数字已经在数组中，返回其索引；如果目标数字不在数组中，
    返回目标数字应该插入的位置，使得插入后数组仍然保持升序。

    :param arr: 升序数组
    :param target: 要查找的目标数字
    :return: 目标数字在数组中的插入位置
    """
    left, right = 0, len(arr) - 1

    while left <= right:
        mid = (left + right) // 2  # 计算中间索引

        if arr[mid] == target:
            return mid  # 找到目标数字，返回其索引
        elif arr[mid] < target:
            left = mid + 1  # 目标在右侧子数组
        else:
            right = mid - 1  # 目标在左侧子数组

    # 如果未找到目标数字，left 的位置即为插入位置
    return left


def      vr_change_c(wavelength,flux,vr):
    wave = np.asarray(wavelength, dtype=np.float64, order='C')
    flux = np.asarray(flux,      dtype=np.float64, order='C')
    wave = np.require(wave, requirements='CW')
    n=wave.size
    wave_vr = np.empty_like(wave)
    #wave = np.require(wave, requirements='CW')
    lib.vr_change_export.argtypes=(
        ctypes.POINTER(ctypes.c_double),
        ctypes.POINTER(ctypes.c_double),
        ctypes.c_int,
        ctypes.c_double,
        ctypes.POINTER(ctypes.c_double)
    )
    lib.vr_change_export.restype=None

    lib.vr_change_export(
        wave.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        flux.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        ctypes.c_int(n),ctypes.c_double(vr),
        wave_vr.ctypes.data_as(ctypes.POINTER(ctypes.c_double))
    )

    return wave_vr

def rot_int_cmj_c(w, s, vsini, eps=0.6, nr=10, ntheta=100, dif = 0.0):
    wave = np.asarray(w, dtype=np.float64, order='C')
    flux = np.asarray(s,      dtype=np.float64, order='C')
    wave = np.require(wave, requirements='CW')
    flux = np.require(flux, requirements='CW')
    n=wave.size
    ns=np.empty_like(wave)

    lib.rot_int_cmj_export.argtypes=(
        ctypes.POINTER(ctypes.c_double),
        ctypes.POINTER(ctypes.c_double),
        ctypes.c_int,
        ctypes.c_double,
        ctypes.c_double,ctypes.c_int,ctypes.c_int,ctypes.c_double,
        ctypes.POINTER(ctypes.c_double)
    )
    lib.rot_int_cmj_export.restype=None

    lib.rot_int_cmj_export(
        wave.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        flux.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        ctypes.c_int(n),ctypes.c_double(vsini),
        ctypes.c_double(eps),ctypes.c_int(nr),ctypes.c_int(ntheta),ctypes.c_double(dif),
        ns.ctypes.data_as(ctypes.POINTER(ctypes.c_double))
    )
    return ns


def get_ck_gk_c(vr,vsini,wavelength_model,flux_model,
              wavelength_H,flux_H,error_H,
              wavelength_K,flux_K,error_K,):
    wave_model = np.asarray(wavelength_model, dtype=np.float64, order='C')
    flux_model = np.asarray(flux_model,      dtype=np.float64, order='C')
    wave_model = np.require(wave_model, requirements='CW')
    flux_model = np.require(flux_model, requirements='CW')
    n_model=wave_model.size

    wave_H = np.asarray(wavelength_H,dtype=np.float64, order='C')
    flux_H = np.asarray(flux_H,      dtype=np.float64, order='C')
    error_H=np.asanyarray(error_H,   dtype=np.float64, order='C')
    wave_H = np.require(wave_H, requirements='CW')
    flux_H = np.require(flux_H, requirements='CW')
    error_H=np.require(error_H, requirements='CW')
    n_H=wave_H.size

    wave_K = np.asarray(wavelength_K,dtype=np.float64, order='C')
    flux_K = np.asarray(flux_K,      dtype=np.float64, order='C')
    error_K=np.asanyarray(error_K,   dtype=np.float64, order='C')
    wave_K = np.require(wave_K, requirements='CW')
    flux_K = np.require(flux_K, requirements='CW')
    error_K=np.require(error_K, requirements='CW')
    n_K=wave_K.size

    #print("原始长度：",n_model,n_H,n_K)

    lib.get_ck_gk_export.argtypes=(
        ctypes.POINTER(ctypes.c_double),
        ctypes.POINTER(ctypes.c_double),
        ctypes.c_size_t, 
        ctypes.POINTER(ctypes.c_double),
        ctypes.POINTER(ctypes.c_double),
        ctypes.POINTER(ctypes.c_double),
        ctypes.c_size_t, 
        ctypes.POINTER(ctypes.c_double),
        ctypes.POINTER(ctypes.c_double),
        ctypes.POINTER(ctypes.c_double),
        ctypes.c_size_t, 
        ctypes.c_double,ctypes.c_double,
        ctypes.POINTER(ctypes.c_double),ctypes.POINTER(ctypes.c_double)
    )
    lib.get_ck_gk_export.restype=None
    ck_out = np.empty(1, dtype=np.float64)
    gk_out = np.empty(1, dtype=np.float64)
    lib.get_ck_gk_export(
        wave_model.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        flux_model.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        ctypes.c_size_t(n_model), 
        wave_H.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        flux_H.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        error_H.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        ctypes.c_size_t(n_H), 
        wave_K.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        flux_K.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        error_K.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        ctypes.c_size_t(n_K), 
        ctypes.c_double(vr),ctypes.c_double(vsini),
        ck_out.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),
        gk_out.ctypes.data_as(ctypes.POINTER(ctypes.c_double)),

    )
    #print('ck结果：',ck_out[0],"Gk 结果:",gk_out[0])
    return vr,vsini,ck_out[0],gk_out[0]

def get_cut(file_path):

    # 设置波长和分辨率参数# Set the wavelength and resolution parameters
    lammin = 0.8  # 最小波长 (μm)
    lammax = 2.5   # 最大波长 (μm)
    R = 45000      # 分辨率


    model_path=file_path
    wave,flux,error=get_txt(model_path)
    #print(wave,len(wave),flux,len(flux))


    # 构造低分辨率的波长网格
    wl, dwl = cg.noise_routines.construct_lam(lammin, lammax, R)

    # 使用 downbin_spec 函数将高分辨率光谱降采样到低分辨率
    flr = cg.downbin_spec(flux, wave, wl, dlam=dwl)

    # 剔除包含 NaN 的部分
    mask = ~np.isnan(wl) 
    wl_clean = wl[mask]
    flr_clean = flr[mask]
    mask = ~np.isnan(flr)
    wl_clean = wl[mask]
    flr_clean = flr[mask]
    wl,flr = wl_clean, flr_clean

    return wl,flr



def parallel_calculation():
    time1 = time.time()
    n=20
    
    # 使用ThreadPoolExecutor创建线程池
    with ThreadPoolExecutor(max_workers=n) as executor:
        # 创建10个任务，第一个参数从1到10
        futures = []
        for i in range(n):
            # 提交任务到线程池
            future = executor.submit(
                get_ck_gk_c, 
                i, 15, wave_model, flux_model, 
                wave_H, flux_H, error_H, 
                wave_K, flux_K, error_K
            )
            futures.append(future)
        
        # 等待所有任务完成
        for future in as_completed(futures):
            try:
                result = future.result()  # 获取计算结果
               # print(f"线程 {futures.index(future)+1} 完成")
            except Exception as e:
                print(f"线程出错: {e}")
    
    time2 = time.time()
    print("并行计算总用时：", time2-time1, ' s',"average: ",(time2-time1)/n,"s")

def excel_and_add_row(file_path, data):
    """
    创建一个Excel文件，并在表格末尾增加一行信息。
    
    参数:
        file_path (str): Excel文件的路径。
        data (list): 要添加到表格末尾的行数据。
    """
    # 加载现有的工作簿
    workbook = load_workbook(file_path)
    # 获取默认的工作表
    sheet = workbook.active
    
    # 在表格末尾添加一行数据
    sheet.append(data)
    
    # 保存工作簿到指定路径
    workbook.save(file_path)
    #print(f"Excel文件已创建并添加了一行数据：{file_path}")

def create_excel_and_add_row(file_path, data):
    """
    创建一个Excel文件，并在表格末尾增加一行信息。
    
    参数:
        file_path (str): Excel文件的路径。
        data (list): 要添加到表格末尾的行数据。
    """
    # 获取目录路径
    dir_path = os.path.dirname(file_path)

    # 如果目录不存在，则创建
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)

    # 创建一个新的工作簿
    workbook = Workbook()
    # 获取默认的工作表
    sheet = workbook.active
    
    # 在表格末尾添加一行数据
    sheet.append(data)
    
    # 保存工作簿到指定路径
    workbook.save(file_path)
    print(f"Excel文件已创建并添加了一行数据：{file_path}")


def extract_values_from_txt(file_path):
    """
    从指定的txt文件中提取teff、logg、meta和alpha的数值。
    
    参数:
        file_path (str): txt文件的路径。
    
    返回:
        dict: 包含teff、logg、meta和alpha的字典。
    """
    # 初始化一个字典来存储提取的数值
    values = {
        "teff": None,
        "logg": None,
        "meta": None,
        "alpha": None
    }
    
    try:
        # 打开文件并逐行读取
        with open(file_path, "r") as file:
            for line in file:
                # 去掉行首和行尾的空白字符
                line = line.strip()
                # 检查是否包含teff
                if "teff" in line:
                    # 提取teff的数值
                    values["teff"] = float(line.split("=")[1].split("K")[0].strip())
                # 检查是否包含logg
                elif "logg" in line:
                    # 提取logg的数值
                    values["logg"] = float(line.split("=")[1].split("log")[0].strip())
                # 检查是否包含meta
                elif "meta" in line:
                    # 提取meta的数值
                    values["meta"] = float(line.split("=")[1].split('(')[0].strip())
                # 检查是否包含alpha
                elif "alpha" in line:
                    # 提取alpha的数值
                    values["alpha"] = float(line.split("=")[1].split('(')[0].strip())
    except FileNotFoundError:
        print(f"文件 {file_path} 未找到。")
    except Exception as e:
        print(f"读取文件时发生错误：{e}")
    
    return values

def vacuum_to_air(wl_Ang):
    '''
        Converts vacuum wavelengths to air wavelengths using the Ciddor 1996 formula.

        :param wl: input vacuum wavelengths
        :type wl: np.array

        :returns: **wl_air** (*np.array*) - the wavelengths converted to air wavelengths

        .. note::

        CA Prieto recommends this as more accurate than the IAU standard.'''

    sigma = (1e4 / wl_Ang) ** 2
    f = 1.0 + 0.05792105 / (238.0185 - sigma) + 0.00167917 / (57.362 - sigma)
    return wl_Ang / f
def air_to_vacuum(wl_Ang):
    ''' convert air wavelength to vacuum wavelength using the Ciddor 1996 formula
    
        when computing sigma, I simply assume vacuum and air wavelength are the same (as sigma are anyway computed from wavelength in um)
    '''
    sigma = (1e4 / wl_Ang)**2
    f = 1.0 + 0.05792105 / (238.0185 - sigma) + 0.00167917 / (57.362 - sigma)
    return wl_Ang * f

def parallel_calculation_1(vr,vsinis,wave_model, flux_model, 
                wave_H, flux_H, error_H, 
                wave_K, flux_K, error_K):
    #time1 = time.time()
    n=12
    
    # 使用ThreadPoolExecutor创建线程池
    with ThreadPoolExecutor(max_workers=n) as executor:
        # 创建10个任务，第一个参数从1到10
        futures = []
        for i in vsinis:
            # 提交任务到线程池
            future = executor.submit(
                get_ck_gk_c, 
                vr, i, wave_model, flux_model, 
                wave_H, flux_H, error_H, 
                wave_K, flux_K, error_K
            )
            futures.append(future)
        
        # 等待所有任务完成
        results = []
        for future in as_completed(futures):
            try:
                result = future.result()  # 获取计算结果
        #print('bbbbbb', result)
                results.append(result)  # 将结果添加到列表中
            except Exception as e:
                 print(f"线程出错: {e}")
        return results  # 返回所有结果
    
   ## time2 = time.time()
    #print("并行计算总用时：", time2-time1, ' s',"average: ",(time2-time1)/n,"s")
def parallel_calculation_2(data_list,
                           wave_model, flux_model,
                           wave_H, flux_H, error_H,
                           wave_K, flux_K, error_K):
    """
    并行处理 data_list 前 16 条记录，
    用 get_ck_gk_c 返回的 (new_vr, new_vsini) 替换 item[8]、item[9]。
    返回修改后的完整 data_list。
    """
    n = 16
    batch = data_list[:n]          # 只处理前 16 条

    def _task(idx, item):
        # 计算新值
        new_vr, new_vsini = get_ck_gk_c(
            item[6], item[7],      # 旧 vr、vsini 仍作为输入
            wave_model, flux_model,
            wave_H, flux_H, error_H,
            wave_K, flux_K, error_K
        )[2:4]
        return idx, new_vr, new_vsini

    with ThreadPoolExecutor(max_workers=min(len(batch), 16)) as exe:
        futures = [exe.submit(_task, i, item) for i, item in enumerate(batch)]

        for fut in as_completed(futures):
            idx, new_vr, new_vsini = fut.result()
            data_list[idx][8] = new_vr
            data_list[idx][9] = new_vsini

    return data_list

if __name__=='__main__':
    H_path='code/part_2/data_for_part2/H_BC.txt'
    K_path='code/part_2/data_for_part2/K_BC.txt'

    #加载观测数据阶段
    wavelength_H,flux_H,error_H=sort(*get_txt(H_path))
    wavelength_K,flux_K,error_K=sort(*get_txt(K_path))

    #最终结果保存路径
    excel_path='code/part_2/outcome_for_part2/ck_gk_vr_vsini_modified_BC_paraller.xlsx'
    #创建excel文件，并添加表头
    data = ["number","file name", "teff", "logg","meta","alpha","vr","vsini","Ck","Gk"]  # 要添加的行数据
    create_excel_and_add_row(excel_path, data)
    number=1

    input_dir = 'models/bt-settle' # 替换为你的文件夹路径
    for file_name in os.listdir(input_dir):
        #time1=time.time()
        if file_name.endswith('.txt'):
            print("正在处理文件：",file_name)
            model_path=os.path.join(input_dir, file_name)
           # wavelength_model,flux_model,aaaaa=get_txt(model_path)
            wavelength_model,flux_model=get_cut(model_path)
            wavelength_model=air_to_vacuum(wavelength_model*1e4)/1e4  #转换为真空波长

            data_list=[]

            for vr in range(-10,10,2):
                for vsinis in range(0,21,5):
                    #print("处理 vr=",vr," vsini=",vsinis)        
                    try:
                        values = extract_values_from_txt(model_path)
                        single_data=[number,file_name,values["teff"],values["logg"],values["meta"],values["alpha"],vr,vsinis,-1,-1]
                        number+=1
                        data_list.append(single_data)

                        if len(data_list)>=16:
                            data_list=parallel_calculation_2(data_list,
                                                             wavelength_model, flux_model,
                                                             wavelength_H, flux_H, error_H,
                                                             wavelength_K, flux_K, error_K)
                            # 批量写入Excel
                            for row in data_list:
                                excel_and_add_row(excel_path, row)
                            data_list=[]

                            try:
                                
                                print("用时",time.time()-time1,' s ',"average:",(time.time()-time1)/16,' s ')
                                time1=time.time()
                            except:
                                time1=time.time()
                            
                                
                    except Exception as e:
                        print("处理出错，跳过该组合。错误信息：", e)
                        continue

                
                    # print("处理vr=",vr)
                    # try:
                    #     result=parallel_calculation_1(vr=vr,vsinis=vsinis,
                    #                     wave_model=wavelength_model,flux_model=flux_model,
                    #                     wave_H=wavelength_H,flux_H=flux_H,error_H=error_H,
                    #                     wave_K=wavelength_K,flux_K=flux_K,error_K=error_K)
                    #     #print("aaaa==========",result)

                    #     for line in result:

                        
                    #     #data = [number,file_name.split('.')[0],file_name.split('_')[2],file_name.split('_')[3],file_name.split('_')[4],file_name.split('_')[5],vr,vsini,ck,gk]  # 要添加的行数据
                    #     #excel_and_add_row(excel_path, data)
                    #         values = extract_values_from_txt(model_path)
                    #         print(values)  # 输出提取的数值
                    #         excel_and_add_row(excel_path, [number,file_name, values["teff"], values["logg"], values["meta"], values["alpha"], line[0], line[1],line[2], line[3]])

                    #         number+=1
                    # except Exception as e:
                    #     print("处理出错，跳过该组合。错误信息：", e)
                    #     continue
        #time2=time.time()
        #print("一份文件用时：",time2-time1,' s ')