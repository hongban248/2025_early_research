#本文件先cut原始的bt-settle模型，然后进行vr，vsini处理，最后插值到观测波段
#本脚本绘制全部的ck，gk图像，经过vr，vsini处理后的结果
#本脚本使用了air_to_vacuum函数将模型从空气波长转换为真空波长，原始数据BC处理了

#尝试这个降分变率
import coronagraph as cg  #这个包还有疑似神秘小bug，需要修改一行代码才能调用
#print(cg.__version__)
import coronagraph as cg
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
from cycler import cycler
import math
from scipy.interpolate import interp1d, CubicSpline
from openpyxl import Workbook,load_workbook
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os


def find_insert_position(arr, target):
    """
    使用二分查找算法查找目标数字在升序数组中的插入位置。
    如果目标数字已经在数组中，返回其索引；如果目标数字不在数组中，
    返回目标数字应该插入的位置，使得插入后数组仍然保持升序。

    :param arr: 升序数组
    :param target: 要查找的目标数字
    :return: 目标数字在数组中的插入位置
    """
    left, right = 0, len(arr) - 1

    while left <= right:
        mid = (left + right) // 2  # 计算中间索引

        if arr[mid] == target:
            return mid  # 找到目标数字，返回其索引
        elif arr[mid] < target:
            left = mid + 1  # 目标在右侧子数组
        else:
            right = mid - 1  # 目标在左侧子数组

    # 如果未找到目标数字，left 的位置即为插入位置
    return left
def interp_1d(x, y, x_new, kind='linear', extrapolate=True):
    """
    一维插值函数

    参数:
        x (array-like): 原始横坐标（一维）
        y (array-like): 原始纵坐标（一维）
        x_new (array-like): 需要插值的新横坐标
        kind (str): 插值类型，可选 'linear', 'nearest', 'zero', 'slinear', 'quadratic', 'cubic', 或 'cubic_spline'
        extrapolate (bool): 是否允许外推（超出原始范围时）

    返回:
        np.ndarray: 插值后的纵坐标
    """
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    x_new = np.asarray(x_new, dtype=float)

    if len(x) != len(y):
        raise ValueError("x 和 y 的长度必须相同")

    if kind == 'cubic_spline':
        spline = CubicSpline(x, y, extrapolate=extrapolate)
        return spline(x_new)
    else:
        f = interp1d(
            x, y,
            kind=kind,
            bounds_error=not extrapolate,
            fill_value='extrapolate' if extrapolate else np.nan
        )
        return f(x_new)

def get_txt(file_path):
    try:
        wavelength, flux, error = np.loadtxt(file_path,
                                     comments='#',      # 跳过所有 # 开头的行
                                     unpack=True)       # 直接拆成三列数组
        return wavelength[find_insert_position(wavelength,0.8):find_insert_position(wavelength,2.5)],flux[find_insert_position(wavelength,0.8):find_insert_position(wavelength,2.5)],error[find_insert_position(wavelength,0.8):find_insert_position(wavelength,2.5)]
    except:
        wavelength, flux = np.loadtxt(file_path,
                                     comments='#',      # 跳过所有 # 开头的行
                           unpack=True)       # 直接拆成2列数组
        wavelength=wavelength/1e4
        e=flux
        return wavelength[find_insert_position(wavelength,0.8):find_insert_position(wavelength,2.5)],flux[find_insert_position(wavelength,0.8):find_insert_position(wavelength,2.5)],e[find_insert_position(wavelength,0.8):find_insert_position(wavelength,2.5)]
def sort(wavelength_1d,spectrum_1d,var_1d):
    wavelength_1d=np.asarray(wavelength_1d)
    spectrum_1d=np.asarray(spectrum_1d)
    var_1d=np.asarray(var_1d)

    idx = np.argsort(wavelength_1d)

    # 2. 重新排序三个数组
    wavelength_sorted = wavelength_1d[idx]
    spectrum_sorted   = spectrum_1d[idx]
    var_sorted        = var_1d[idx]

    return wavelength_sorted, spectrum_sorted, var_sorted


def get_cut(file_path):

    # 设置波长和分辨率参数# Set the wavelength and resolution parameters
    lammin = 0.8  # 最小波长 (μm)
    lammax = 2.5   # 最大波长 (μm)
    R = 45000      # 分辨率


    model_path=file_path
    wave,flux,error=get_txt(model_path)
    #print(wave,len(wave),flux,len(flux))


    # 构造低分辨率的波长网格
    wl, dwl = cg.noise_routines.construct_lam(lammin, lammax, R)

    # 使用 downbin_spec 函数将高分辨率光谱降采样到低分辨率
    flr = cg.downbin_spec(flux, wave, wl, dlam=dwl)

    # 剔除包含 NaN 的部分
    mask = ~np.isnan(wl) 
    wl_clean = wl[mask]
    flr_clean = flr[mask]
    mask = ~np.isnan(flr)
    wl_clean = wl[mask]
    flr_clean = flr[mask]
    wl,flr = wl_clean, flr_clean

    return wl,flr


       
def interp_spec(wavelength_K,flux_K,error_K,wavelength1,flux1,error1):
    # 对高分辨率光谱进行插值到K上波段的波长点
    left_index=find_insert_position(wavelength1,wavelength_K[0])
    print("left_index:",left_index)
    right_index=find_insert_position(wavelength1,wavelength_K[-1])

    wavelength_temp=[]
    flux_temp=[]
    error_temp=[]
    # for i in range(left_index,right_index):
    #     wavelength_temp.append(wavelength_K[i-left_index])
    #     flux_new=interp_1d(wavelength_K,flux_K,wavelength1[i],kind='linear',extrapolate=True)
    #     error_new=interp_1d(wavelength_K,error_K,error1[i],kind='linear',extrapolate=True)
    #     flux_temp.append(flux_new)
    #     error_temp.append(error_new)
    for i in range(len(wavelength_K)):
        wavelength_temp.append(wavelength_K[i])

        flux_new=interp_1d(wavelength1[left_index:right_index],flux1[left_index:right_index],wavelength_K[i],kind='linear',extrapolate=True)
        error_new=interp_1d(wavelength1[left_index:right_index],error1[left_index:right_index],wavelength_K[i],kind='linear',extrapolate=True)

        flux_temp.append(flux_new)
        error_temp.append(error_new)

    # wavelength=wavelength1[0:left_index].tolist()+wavelength_temp+wavelength1[right_index:].tolist()
    # flux=flux1[0:left_index].tolist()+flux_temp+flux1[right_index:].tolist()
    # error=error1[0:left_index].tolist()+error_temp+error1[right_index:].tolist()
    #return np.array(wavelength),np.array(flux),np.array(error)
    return wavelength_temp,flux_temp,error_temp



def rot_int_cmj(w, s, vsini, eps=0.6, nr=10, ntheta=100, dif = 0.0):
    '''
    A routine to quickly rotationally broaden a spectrum in linear time.

    INPUTS:
    s - input spectrum

    w - wavelength scale of the input spectrum
    
    vsini (km/s) - projected rotational velocity
    
    OUTPUT:
    ns - a rotationally broadened spectrum on the wavelength scale w

    OPTIONAL INPUTS:
    eps (default = 0.6) - the coefficient of the limb darkening law
    
    nr (default = 10) - the number of radial bins on the projected disk
    
    ntheta (default = 100) - the number of azimuthal bins in the largest radial annulus
                            note: the number of bins at each r is int(r*ntheta) where r < 1
    
    dif (default = 0) - the differential rotation coefficient, applied according to the law
    Omeg(th)/Omeg(eq) = (1 - dif/2 - (dif/2) cos(2 th)). Dif = .675 nicely reproduces the law
    proposed by Smith, 1994, A&A, Vol. 287, p. 523-534, to unify WTTS and CTTS. Dif = .23 is
    similar to observed solar differential rotation. Note: the th in the above expression is
    the stellar co-latitude, not the same as the integration variable used below. This is a
    disk integration routine.
    
    
    
    Reference:
    RNAAS: Carvalho & Johns-Krull (2023)
    '''

    ns = np.copy(s)*0.0
    tarea = 0.0
    dr = 1./nr
    for j in range(0, nr):
        r = dr/2.0 + j*dr
        area = ((r + dr/2.0)**2 - (r - dr/2.0)**2)/int(ntheta*r) * (1.0 - eps + eps*np.cos(np.arcsin(r)))
        for k in range(0,int(ntheta*r)):
            th = np.pi/int(ntheta*r) + k * 2.0*np.pi/int(ntheta*r)
            if dif != 0:
                vl = vsini * r * np.sin(th) * (1.0 - dif/2.0 - dif/2.0*np.cos(2.0*np.arccos(r*np.cos(th))))
                ns += area * np.interp(w + w*vl/2.9979e5, w, s)
                tarea += area
            else:
                vl = r * vsini * np.sin(th)
                ns += area * np.interp(w + w*vl/2.9979e5, w, s)
                tarea += area
          
    return ns/tarea

def vr_change(wavelength,flux,vr):
    #对光谱进行径向速度平移
    wavelength_vr=wavelength*(1+vr/299792.458)
    return wavelength_vr

def get_Ck(wave,flux,wave_K,flux_K,error_K,wave_H,flux_H,error_H,wave_model_H,flux_model_H):

    #left_index=0
    # #print("left_index",left_index)
    # right_index=find_insert_position(wave,wave_K[-1])

    '''
    sum1 = np.sum( flux * flux_K / err_K**2 ) + np.sum( flux * flux_H / err_H**2 )
    sum2 = np.sum( flux**2 / err_K**2 ) + np.sum( flux**2 / err**2 )#有问题

    sum1 = np.sum( flux_model_K * flux_K / err_K**2 ) + np.sum( flux_model_H * flux_H / err_H**2 )
    sum2 = np.sum( flux_model_K**2 / err_K**2 ) + np.sum( flux_model_H**2 / err_H**2 )

    '''

    sum1=0
    sum2=0
    for i in range(0,len(wave_K)):
        #print(i,left_index+i,wave_K[i],wave[left_index+i])
        #break
        sum1=sum1+flux[i]*flux_K[i]/(error_K[i]**2)
        sum2=sum2+flux[i]**2/(error_K[i]**2)

    for i in range(0,len(wave_H)):
        sum1=sum1+flux_model_H[i]*flux_H[i]/(error_H[i]**2)
        sum2=sum2+flux_model_H[i]**2/(error_H[i]**2)

    Ck=sum1/sum2
    #Ck=sum(wave[left_index+i]*wave_K[i]/(error_K[i]**2))/sum(wave[left_index+i]**2/(error_K[i]**2))
    print('Ck:',Ck)
    return Ck

def get_Gk(wave,flux,wave_K,flux_K,error_K,Ck,wave_H,flux_H,error_H,wave_model_H,flux_model_H):

    # left_index=0
    # #print("left_index",left_index)
    # right_index=find_insert_position(wave,wave_K[-1])

    sum1=0
    
    for i in range(0,len(wave_K)):
        #print(i,left_index+i,wave_K[i],wave[left_index+i])
        #break
        sum1=sum1+(flux_K[i]-Ck*flux[i])**2/(error_K[i]**2)

    for i in range(0,len(wave_H)):
        sum1=sum1+(flux_H[i]-Ck*flux_model_H[i])**2/(error_H[i]**2)

    Gk=sum1
    #Ck=sum(wave[left_index+i]*wave_K[i]/(error_K[i]**2))/sum(wave[left_index+i]**2/(error_K[i]**2))
    print('Gk:',Gk)
    return Gk
    
def get_ck_gk(vr,vsini,wavelength_model,flux_model,
              wavelength_H,flux_H,error_H,
              wavelength_K,flux_K,error_K,
              save_model_vr_vsini_H=False,save_model_vr_vsini_K=False,original_model_name='model.txt'):

    wavelength_model_vr=vr_change(wavelength_model,flux_model,vr)
    flux_model_vr_vsini=rot_int_cmj(wavelength_model_vr,flux_model,vsini)

    wavelength_model_H,flux_model_H,error_model_H=interp_spec(wavelength_H,flux_H,error_H,wavelength_model_vr,flux_model_vr_vsini,np.zeros(len(flux_model_vr_vsini)))
    wavelength_model_K,flux_model_K,error_model_K=interp_spec(wavelength_K,flux_K,error_K,wavelength_model_vr,flux_model_vr_vsini,np.zeros(len(flux_model_vr_vsini)))
    
    #print(wavelength_H[0],wavelength_model_H[0])
    #这里选择是否保存中间过程'
    if save_model_vr_vsini_H:
        os.makedirs('models/bt-settle_H_vr_vsini', exist_ok=True)
        save_path_H='models/bt-settle_H_vr_vsini/'+original_model_name.split('/')[-1]+'_vr'+str(vr)+'_vsini'+str(vsini)+'.txt'
        np.savetxt(save_path_H,np.column_stack((wavelength_model_H,flux_model_H,error_model_H)),fmt='%.6f %.6e %.6e',header='# Wavelength(um)   Flux   Error')
        print("保存文件到：",save_path_H)

    if save_model_vr_vsini_K:
        os.makedirs('models/bt-settle_K_vr_vsini', exist_ok=True)
        save_path_K='models/bt-settle_K_vr_vsini/'+original_model_name.split('/')[-1]+'_vr'+str(vr)+'_vsini'+str(vsini)+'.txt'
        np.savetxt(save_path_K,np.column_stack((wavelength_model_K,flux_model_K,error_model_K)),fmt='%.6f %.6e %.6e',header='# Wavelength(um)   Flux   Error')
        print("保存文件到：",save_path_K)

    #下面计算ck，gk
    ck=get_Ck(wavelength_model_K,flux_model_K,wavelength_K,flux_K,error_K,wavelength_H,flux_H,error_H,wavelength_model_H,flux_model_H)
    gk=get_Gk(wavelength_model_K,flux_model_K,wavelength_K,flux_K,error_K,ck,wavelength_H,flux_H,error_H,wavelength_model_H,flux_model_H)

    return ck,gk


def excel_and_add_row(file_path, data):
    """
    创建一个Excel文件，并在表格末尾增加一行信息。
    
    参数:
        file_path (str): Excel文件的路径。
        data (list): 要添加到表格末尾的行数据。
    """
    # 加载现有的工作簿
    workbook = load_workbook(file_path)
    # 获取默认的工作表
    sheet = workbook.active
    
    # 在表格末尾添加一行数据
    sheet.append(data)
    
    # 保存工作簿到指定路径
    workbook.save(file_path)
    print(f"Excel文件已创建并添加了一行数据：{file_path}")

def create_excel_and_add_row(file_path, data):
    """
    创建一个Excel文件，并在表格末尾增加一行信息。
    
    参数:
        file_path (str): Excel文件的路径。
        data (list): 要添加到表格末尾的行数据。
    """
    # 创建一个新的工作簿
    workbook = Workbook()
    # 获取默认的工作表
    sheet = workbook.active
    
    # 在表格末尾添加一行数据
    sheet.append(data)
    
    # 保存工作簿到指定路径
    workbook.save(file_path)
    print(f"Excel文件已创建并添加了一行数据：{file_path}")


def extract_values_from_txt(file_path):
    """
    从指定的txt文件中提取teff、logg、meta和alpha的数值。
    
    参数:
        file_path (str): txt文件的路径。
    
    返回:
        dict: 包含teff、logg、meta和alpha的字典。
    """
    # 初始化一个字典来存储提取的数值
    values = {
        "teff": None,
        "logg": None,
        "meta": None,
        "alpha": None
    }
    
    try:
        # 打开文件并逐行读取
        with open(file_path, "r") as file:
            for line in file:
                # 去掉行首和行尾的空白字符
                line = line.strip()
                # 检查是否包含teff
                if "teff" in line:
                    # 提取teff的数值
                    values["teff"] = float(line.split("=")[1].split("K")[0].strip())
                # 检查是否包含logg
                elif "logg" in line:
                    # 提取logg的数值
                    values["logg"] = float(line.split("=")[1].split("log")[0].strip())
                # 检查是否包含meta
                elif "meta" in line:
                    # 提取meta的数值
                    values["meta"] = float(line.split("=")[1].split('(')[0].strip())
                # 检查是否包含alpha
                elif "alpha" in line:
                    # 提取alpha的数值
                    values["alpha"] = float(line.split("=")[1].split('(')[0].strip())
    except FileNotFoundError:
        print(f"文件 {file_path} 未找到。")
    except Exception as e:
        print(f"读取文件时发生错误：{e}")
    
    return values

def vacuum_to_air(wl_Ang):
    '''
        Converts vacuum wavelengths to air wavelengths using the Ciddor 1996 formula.

        :param wl: input vacuum wavelengths
        :type wl: np.array

        :returns: **wl_air** (*np.array*) - the wavelengths converted to air wavelengths

        .. note::

        CA Prieto recommends this as more accurate than the IAU standard.'''

    sigma = (1e4 / wl_Ang) ** 2
    f = 1.0 + 0.05792105 / (238.0185 - sigma) + 0.00167917 / (57.362 - sigma)
    return wl_Ang / f
def air_to_vacuum(wl_Ang):
    ''' convert air wavelength to vacuum wavelength using the Ciddor 1996 formula
    
        when computing sigma, I simply assume vacuum and air wavelength are the same (as sigma are anyway computed from wavelength in um)
    '''
    sigma = (1e4 / wl_Ang)**2
    f = 1.0 + 0.05792105 / (238.0185 - sigma) + 0.00167917 / (57.362 - sigma)
    return wl_Ang * f


if __name__ == "__main__":
    #文件路径设置
    H_path='code/part_2/data_for_part2/H_BC.txt'
    K_path='code/part_2/data_for_part2/K_BC.txt'

    #加载观测数据阶段
    wavelength_H,flux_H,error_H=sort(*get_txt(H_path))
    wavelength_K,flux_K,error_K=sort(*get_txt(K_path))

    #最终结果保存路径
    excel_path='code/part_2/outcome_for_part2/ck_gk_vr_vsini_modified.xlsx'
    #创建excel文件，并添加表头
    data = ["number","file name", "teff", "logg","meta","alpha","vr","vsini","Ck","Gk"]  # 要添加的行数据
    create_excel_and_add_row(excel_path, data)
    number=1

    input_dir = 'models/bt-settle' # 替换为你的文件夹路径
    for file_name in os.listdir(input_dir):
        if file_name.endswith('.txt'):
            print("正在处理文件：",file_name)
            model_path=os.path.join(input_dir, file_name)
            wavelength_model,flux_model,aaaaa=get_txt(model_path)
            wavelength_model,flux_model=get_cut(model_path)
            wavelength_model=air_to_vacuum(wavelength_model*1e4)/1e4  #转换为真空波长

            for vr in range(-10,10,2):
                for vsini in range(0,20,4):
                    print("处理vr=",vr,"vsini=",vsini)
                    try:
                        ck,gk=get_ck_gk(vr=vr,vsini=vsini,
                                        wavelength_model=wavelength_model,flux_model=flux_model,
                                        wavelength_H=wavelength_H,flux_H=flux_H,error_H=error_H,
                                        wavelength_K=wavelength_K,flux_K=flux_K,error_K=error_K,
                                        save_model_vr_vsini_H=False,save_model_vr_vsini_K=False,original_model_name=file_name)
                        
                        #data = [number,file_name.split('.')[0],file_name.split('_')[2],file_name.split('_')[3],file_name.split('_')[4],file_name.split('_')[5],vr,vsini,ck,gk]  # 要添加的行数据
                        #excel_and_add_row(excel_path, data)
                        values = extract_values_from_txt(model_path)
                        print(values)  # 输出提取的数值
                        excel_and_add_row(excel_path, [number,file_name, values["teff"], values["logg"], values["meta"], values["alpha"], vr, vsini, ck, gk])

                        number+=1
                    except Exception as e:
                        print("处理出错，跳过该组合。错误信息：", e)
                        continue